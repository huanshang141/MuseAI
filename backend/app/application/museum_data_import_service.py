"""Strict, idempotent museum-data import for the mini-program runtime.

The parser validates the complete workbook/CSV pair before a database session is
needed. The importer then stages changed exhibits as inactive, rebuilds their
RAG index, and only exposes successfully indexed exhibits to the mini-program.
Rows omitted from a later import are deliberately left untouched unless the
caller explicitly enables authoritative snapshot cleanup.
"""

from __future__ import annotations

import csv
import json
import math
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol
from xml.etree.ElementTree import ParseError as XMLParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.content_source import ContentMetadata, ContentSource
from app.application.exhibit_images import normalize_external_image_url
from app.infra.postgres.models import Exhibit, Hall

HALL_HEADERS = (
    "source_record_id",
    "slug",
    "name",
    "description",
    "floor",
    "estimated_duration_minutes",
    "display_order",
    "is_active",
    "suggested_questions",
)
EXHIBIT_HEADERS = (
    "source_record_id",
    "name",
    "description",
    "hall",
    "floor",
    "category",
    "era",
    "importance",
    "estimated_visit_time",
    "display_order",
    "location_x",
    "location_y",
    "is_active",
    "suggested_questions",
)
EXHIBIT_OPTIONAL_HEADERS = ("image_url",)
SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
SOURCE_NAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{0,99}$")
EXHIBIT_ID_NAMESPACE = uuid.UUID("d7864da1-b507-5fb6-bb2d-c8de1dadbf58")
MAX_ACTIVE_EXHIBITS = 2_000
SUGGESTION_MAX_LENGTH = 120


class MuseumDataValidationError(ValueError):
    """Raised after collecting all input validation problems."""

    def __init__(self, issues: list[str]):
        self.issues = issues
        super().__init__("Museum data validation failed:\n- " + "\n- ".join(issues))


class MuseumDataIndexingError(RuntimeError):
    """Raised when DB staging succeeded but one or more RAG updates failed."""

    def __init__(self, failures: list[str], summary: ImportSummary):
        self.failures = failures
        self.summary = summary
        super().__init__("Museum data RAG update was incomplete: " + "; ".join(failures))


class IndexingService(Protocol):
    async def index_source(self, source: ContentSource, max_concurrency: int = 10) -> int: ...

    async def delete_source(self, source_id: str, source_type: str | None = None) -> dict[str, Any]: ...


@dataclass(frozen=True)
class HallImportRow:
    source_record_id: str
    slug: str
    name: str
    description: str
    floor: int | None
    estimated_duration_minutes: int
    display_order: int
    is_active: bool
    suggested_questions: list[str]


@dataclass(frozen=True)
class ExhibitImportRow:
    source_record_id: str
    name: str
    description: str
    hall: str
    floor: int | None
    category: str | None
    era: str | None
    importance: int
    estimated_visit_time: int | None
    display_order: int
    location_x: float | None
    location_y: float | None
    is_active: bool
    suggested_questions: list[str]
    image_url: str | None = None
    image_url_present: bool = False


@dataclass(frozen=True)
class MuseumDataset:
    halls: list[HallImportRow]
    exhibits: list[ExhibitImportRow]


@dataclass
class ImportSummary:
    source_name: str
    dry_run: bool
    halls_seen: int
    exhibits_seen: int
    authoritative: bool = False
    halls_created: int = 0
    halls_updated: int = 0
    halls_deactivated: int = 0
    exhibits_created: int = 0
    exhibits_updated: int = 0
    exhibits_indexed: int = 0
    exhibits_deactivated: int = 0
    unchanged: int = 0
    halls_planned_deactivation: int = 0
    exhibits_planned_deactivation: int = 0
    authoritative_cleanup_deferred: bool = False
    pending_index: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_name": self.source_name,
            "dry_run": self.dry_run,
            "halls_seen": self.halls_seen,
            "exhibits_seen": self.exhibits_seen,
            "authoritative": self.authoritative,
            "halls_created": self.halls_created,
            "halls_updated": self.halls_updated,
            "halls_deactivated": self.halls_deactivated,
            "exhibits_created": self.exhibits_created,
            "exhibits_updated": self.exhibits_updated,
            "exhibits_indexed": self.exhibits_indexed,
            "exhibits_deactivated": self.exhibits_deactivated,
            "unchanged": self.unchanged,
            "halls_planned_deactivation": self.halls_planned_deactivation,
            "exhibits_planned_deactivation": self.exhibits_planned_deactivation,
            "authoritative_cleanup_deferred": self.authoritative_cleanup_deferred,
            "pending_index": self.pending_index,
        }


def deterministic_exhibit_id(source_name: str, source_record_id: str) -> str:
    return str(uuid.uuid5(EXHIBIT_ID_NAMESPACE, f"{source_name}:exhibit:{source_record_id}"))


def validate_source_name(source_name: str) -> str:
    normalized = source_name.strip()
    if not SOURCE_NAME_RE.fullmatch(normalized):
        raise MuseumDataValidationError(["source_name must be 1-100 ASCII letters, digits, '.', '_', ':' or '-'"])
    return normalized


def load_museum_dataset(input_path: str | Path) -> MuseumDataset:
    """Load and fully validate either one .xlsx file or a CSV-pair directory."""
    path = Path(input_path)
    if path.is_file() and path.suffix.lower() == ".xlsx":
        raw_halls, raw_exhibits = _read_xlsx(path)
    elif path.is_dir():
        raw_halls, raw_exhibits = _read_csv_pair(path)
    else:
        raise MuseumDataValidationError(
            ["input must be a .xlsx file or a directory containing halls.csv and exhibits.csv"]
        )
    return _validate_rows(raw_halls, raw_exhibits)


def _read_csv_pair(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    required = {"halls.csv", "exhibits.csv"}
    names = {item.name for item in path.iterdir() if item.is_file()}
    missing = sorted(required - names)
    if missing:
        raise MuseumDataValidationError([f"CSV directory is missing: {', '.join(missing)}"])
    return (
        _read_csv_table(path / "halls.csv", HALL_HEADERS, "halls.csv"),
        _read_csv_table(
            path / "exhibits.csv",
            EXHIBIT_HEADERS,
            "exhibits.csv",
            optional_headers=EXHIBIT_OPTIONAL_HEADERS,
        ),
    )


def _read_csv_table(
    path: Path,
    expected_headers: tuple[str, ...],
    label: str,
    *,
    optional_headers: tuple[str, ...] = (),
) -> list[dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.reader(file, strict=True)
            try:
                rows = list(reader)
            except csv.Error as exc:
                raise MuseumDataValidationError([f"{label} is malformed near line {reader.line_num}: {exc}"]) from None
    except UnicodeDecodeError as exc:
        raise MuseumDataValidationError(
            [f"{label} must be UTF-8 encoded (decode failed near byte {exc.start})"]
        ) from None
    if not rows:
        raise MuseumDataValidationError([f"{label} is empty"])
    return _table_to_dicts(
        rows[0],
        rows[1:],
        expected_headers,
        label,
        optional_headers=optional_headers,
    )


def _read_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        expected = {"halls", "exhibits"}
        actual = set(workbook.sheetnames)
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        issues = []
        if missing:
            issues.append(f"workbook is missing sheets: {', '.join(missing)}")
        if extra:
            issues.append(f"workbook has unexpected sheets: {', '.join(extra)}")
        if issues:
            raise MuseumDataValidationError(issues)
        parsed: dict[str, list[dict[str, Any]]] = {}
        for sheet_name, headers in (("halls", HALL_HEADERS), ("exhibits", EXHIBIT_HEADERS)):
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                raise MuseumDataValidationError([f"sheet '{sheet_name}' is empty"])
            parsed[sheet_name] = _table_to_dicts(
                rows[0],
                rows[1:],
                headers,
                f"sheet '{sheet_name}'",
                optional_headers=(EXHIBIT_OPTIONAL_HEADERS if sheet_name == "exhibits" else ()),
            )
        return parsed["halls"], parsed["exhibits"]
    except MuseumDataValidationError:
        raise
    except (BadZipFile, InvalidFileException, KeyError, XMLParseError, ValueError) as exc:
        raise MuseumDataValidationError([f"workbook is not a valid .xlsx file: {exc}"]) from None
    finally:
        if workbook is not None:
            workbook.close()


def _table_to_dicts(
    raw_headers: Any,
    raw_rows: Any,
    expected_headers: tuple[str, ...],
    label: str,
    *,
    optional_headers: tuple[str, ...] = (),
) -> list[dict[str, Any]]:
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    issues: list[str] = []
    if len(headers) != len(set(headers)):
        issues.append(f"{label} contains duplicate headers")
    missing = sorted(set(expected_headers) - set(headers))
    extra = sorted(set(headers) - set(expected_headers) - set(optional_headers))
    if missing:
        issues.append(f"{label} is missing headers: {', '.join(missing)}")
    if extra:
        issues.append(f"{label} has unexpected headers: {', '.join(extra)}")
    if issues:
        raise MuseumDataValidationError(issues)
    result: list[dict[str, Any]] = []
    for row_number, raw_row in enumerate(raw_rows, start=2):
        cells = list(raw_row)
        if not any(value not in (None, "") for value in cells):
            continue
        if len(cells) > len(headers) and any(value not in (None, "") for value in cells[len(headers) :]):
            raise MuseumDataValidationError([f"{label} row {row_number} has values beyond the header columns"])
        cells = cells[: len(headers)]
        cells.extend([None] * (len(headers) - len(cells)))
        result.append({"__row__": row_number, **dict(zip(headers, cells, strict=True))})
    return result


def _validate_rows(raw_halls: list[dict[str, Any]], raw_exhibits: list[dict[str, Any]]) -> MuseumDataset:
    issues: list[str] = []
    halls: list[HallImportRow] = []
    exhibits: list[ExhibitImportRow] = []
    for raw in raw_halls:
        try:
            halls.append(_parse_hall(raw))
        except ValueError as exc:
            issues.append(f"halls row {raw['__row__']}: {exc}")
    for raw in raw_exhibits:
        try:
            exhibits.append(_parse_exhibit(raw))
        except ValueError as exc:
            issues.append(f"exhibits row {raw['__row__']}: {exc}")

    hall_slugs = [row.slug for row in halls]
    hall_names = [row.name for row in halls]
    hall_source_ids = [row.source_record_id for row in halls]
    exhibit_source_ids = [row.source_record_id for row in exhibits]
    issues.extend(_duplicate_issues(hall_slugs, "duplicate hall slug"))
    issues.extend(_duplicate_issues(hall_names, "duplicate hall name"))
    issues.extend(_duplicate_issues(hall_source_ids, "duplicate hall source_record_id"))
    issues.extend(_duplicate_issues(exhibit_source_ids, "duplicate exhibit source_record_id"))
    if sum(1 for row in halls if row.is_active) > 9:
        issues.append("at most 9 active halls can be imported")
    if sum(1 for row in exhibits if row.is_active) > MAX_ACTIVE_EXHIBITS:
        issues.append(f"at most {MAX_ACTIVE_EXHIBITS} active exhibits can be imported")
    hall_map = {row.slug: row for row in halls}
    for row in exhibits:
        if row.hall not in hall_map:
            issues.append(f"exhibit source_record_id '{row.source_record_id}' references unknown hall '{row.hall}'")
        elif row.is_active and not hall_map[row.hall].is_active:
            issues.append(
                f"active exhibit source_record_id '{row.source_record_id}' references inactive hall '{row.hall}'"
            )
    if not halls:
        issues.append("at least one hall row is required")
    if issues:
        raise MuseumDataValidationError(issues)
    return MuseumDataset(halls=halls, exhibits=exhibits)


def _parse_hall(raw: dict[str, Any]) -> HallImportRow:
    slug = _required_text(raw, "slug", 100)
    if not SLUG_RE.fullmatch(slug):
        raise ValueError("slug must use lowercase letters/digits separated by single hyphens")
    return HallImportRow(
        source_record_id=_required_text(raw, "source_record_id", 255),
        slug=slug,
        name=_required_text(raw, "name", 255),
        description=_required_text(raw, "description", 20_000),
        floor=_optional_int(raw, "floor", minimum=-10, maximum=200),
        estimated_duration_minutes=_required_int(raw, "estimated_duration_minutes", minimum=0, maximum=480),
        display_order=_required_int(raw, "display_order", minimum=0, maximum=1_000_000),
        is_active=_required_bool(raw, "is_active"),
        suggested_questions=_questions(raw.get("suggested_questions"), "suggested_questions"),
    )


def _parse_exhibit(raw: dict[str, Any]) -> ExhibitImportRow:
    return ExhibitImportRow(
        source_record_id=_required_text(raw, "source_record_id", 255),
        name=_required_text(raw, "name", 255),
        description=_required_text(raw, "description", 50_000),
        hall=_required_text(raw, "hall", 100),
        floor=_optional_int(raw, "floor", minimum=-10, maximum=200),
        category=_optional_text(raw, "category", 100),
        era=_optional_text(raw, "era", 100),
        importance=_required_int(raw, "importance", minimum=0, maximum=100),
        estimated_visit_time=_optional_int(raw, "estimated_visit_time", minimum=1, maximum=86_400),
        display_order=_required_int(raw, "display_order", minimum=0, maximum=1_000_000),
        location_x=_optional_float(raw, "location_x"),
        location_y=_optional_float(raw, "location_y"),
        is_active=_required_bool(raw, "is_active"),
        suggested_questions=_questions(raw.get("suggested_questions"), "suggested_questions"),
        image_url=_optional_image_url(raw, "image_url"),
        image_url_present="image_url" in raw,
    )


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _required_text(raw: dict[str, Any], key: str, maximum: int) -> str:
    value = _clean_text(raw.get(key))
    if not value:
        raise ValueError(f"{key} is required")
    if len(value) > maximum:
        raise ValueError(f"{key} exceeds {maximum} characters")
    return value


def _optional_text(raw: dict[str, Any], key: str, maximum: int) -> str | None:
    value = _clean_text(raw.get(key))
    if not value:
        return None
    if len(value) > maximum:
        raise ValueError(f"{key} exceeds {maximum} characters")
    return value


def _optional_image_url(raw: dict[str, Any], key: str) -> str | None:
    value = _optional_text(raw, key, 2048)
    return normalize_external_image_url(value)


def _required_int(raw: dict[str, Any], key: str, *, minimum: int, maximum: int) -> int:
    value = _optional_int(raw, key, minimum=minimum, maximum=maximum)
    if value is None:
        raise ValueError(f"{key} is required")
    return value


def _optional_int(raw: dict[str, Any], key: str, *, minimum: int, maximum: int) -> int | None:
    value = raw.get(key)
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{key} must be an integer")
    try:
        parsed_float = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{key} must be an integer") from exc
    if not parsed_float.is_integer():
        raise ValueError(f"{key} must be an integer")
    parsed = int(parsed_float)
    if not minimum <= parsed <= maximum:
        raise ValueError(f"{key} must be between {minimum} and {maximum}")
    return parsed


def _optional_float(raw: dict[str, Any], key: str) -> float | None:
    value = raw.get(key)
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{key} must be a finite number")
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{key} must be a finite number") from exc
    if not math.isfinite(parsed) or abs(parsed) > 1_000_000:
        raise ValueError(f"{key} must be a finite number between -1000000 and 1000000")
    return parsed


def _required_bool(raw: dict[str, Any], key: str) -> bool:
    value = raw.get(key)
    if isinstance(value, bool):
        return value
    normalized = _clean_text(value).lower()
    if normalized in {"true", "1", "yes", "是"}:
        return True
    if normalized in {"false", "0", "no", "否"}:
        return False
    raise ValueError(f"{key} must be true/false, 1/0, yes/no or 是/否")


def _questions(value: Any, key: str) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{key} contains invalid JSON") from exc
        else:
            parsed = [part.strip() for part in stripped.split("|") if part.strip()]
    elif isinstance(value, (list, tuple)):
        parsed = list(value)
    else:
        raise ValueError(f"{key} must be a JSON list or | separated text")
    if not isinstance(parsed, list) or any(not isinstance(item, str) for item in parsed):
        raise ValueError(f"{key} must contain only strings")
    normalized = [item.strip() for item in parsed if item.strip()]
    if len(normalized) > 6:
        raise ValueError(f"{key} supports at most 6 questions")
    if any(len(item) > SUGGESTION_MAX_LENGTH for item in normalized):
        raise ValueError(f"{key} question exceeds {SUGGESTION_MAX_LENGTH} characters")
    return normalized


def _duplicate_issues(values: list[str], label: str) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return [f"{label}: '{value}'" for value in sorted(duplicates)]


class MuseumDataImportService:
    def __init__(
        self,
        session: AsyncSession | None = None,
        indexing_service: IndexingService | None = None,
    ):
        self.session = session
        self.indexing_service = indexing_service

    async def import_dataset(
        self,
        dataset: MuseumDataset,
        *,
        source_name: str,
        dry_run: bool = False,
        authoritative: bool = False,
    ) -> ImportSummary:
        source_name = validate_source_name(source_name)
        summary = ImportSummary(
            source_name=source_name,
            dry_run=dry_run,
            halls_seen=len(dataset.halls),
            exhibits_seen=len(dataset.exhibits),
            authoritative=authoritative,
        )
        summary.halls_planned_deactivation = sum(1 for row in dataset.halls if not row.is_active)
        summary.exhibits_planned_deactivation = sum(1 for row in dataset.exhibits if not row.is_active)
        if dry_run:
            summary.authoritative_cleanup_deferred = authoritative
            summary.pending_index = [row.source_record_id for row in dataset.exhibits if row.is_active]
            return summary
        if self.session is None:
            raise RuntimeError("A database session is required unless dry_run=True")

        hall_by_slug, hall_by_source, hall_by_name = await self._existing_halls(dataset, source_name)
        exhibit_by_source = await self._existing_exhibits(dataset, source_name)
        authoritative_halls: list[Hall] = []
        authoritative_exhibits: list[Exhibit] = []
        if authoritative:
            authoritative_halls, authoritative_exhibits = await self._authoritative_targets(dataset, source_name)
            for hall in authoritative_halls:
                hall_by_slug.setdefault(hall.slug, hall)

            target_exhibit_ids = {model.id for model in authoritative_exhibits}
            incoming_relocated_or_inactive_ids = {
                model.id
                for row in dataset.exhibits
                if (model := exhibit_by_source.get(row.source_record_id)) is not None
                and (not row.is_active or model.hall != row.hall)
            }
            candidate_hall_slugs = {model.slug for model in authoritative_halls}
            protected_hall_slugs: set[str] = set()
            if candidate_hall_slugs:
                active_rows = (
                    await self.session.execute(
                        select(Exhibit.id, Exhibit.hall).where(
                            Exhibit.is_active.is_(True),
                            Exhibit.hall.in_(candidate_hall_slugs),
                        )
                    )
                ).all()
                protected_hall_slugs = {
                    str(hall)
                    for exhibit_id, hall in active_rows
                    if exhibit_id not in target_exhibit_ids and exhibit_id not in incoming_relocated_or_inactive_ids
                }
            authoritative_halls = [model for model in authoritative_halls if model.slug not in protected_hall_slugs]
            summary.halls_planned_deactivation += len(authoritative_halls)
            summary.exhibits_planned_deactivation += len(authoritative_exhibits)

        collision_issues: list[str] = []
        existing_active_halls = set(
            (await self.session.execute(select(Hall.slug).where(Hall.is_active.is_(True)))).scalars()
        )
        imported_hall_slugs = {row.slug for row in dataset.halls}
        authoritative_hall_slugs = {model.slug for model in authoritative_halls}
        resulting_active_halls = (existing_active_halls - imported_hall_slugs - authoritative_hall_slugs) | {
            row.slug for row in dataset.halls if row.is_active
        }
        if len(resulting_active_halls) > 9:
            collision_issues.append(
                "import would leave more than 9 active halls in the database; "
                "explicitly include and deactivate obsolete halls first"
            )
        existing_active_exhibit_ids = set(
            (await self.session.execute(select(Exhibit.id).where(Exhibit.is_active.is_(True)))).scalars()
        )
        imported_existing_exhibit_ids = {
            model.id for row in dataset.exhibits if (model := exhibit_by_source.get(row.source_record_id)) is not None
        }
        authoritative_exhibit_ids = {model.id for model in authoritative_exhibits}
        imported_active_exhibit_ids = {
            (
                model.id
                if (model := exhibit_by_source.get(row.source_record_id)) is not None
                else deterministic_exhibit_id(source_name, row.source_record_id)
            )
            for row in dataset.exhibits
            if row.is_active
        }
        resulting_active_exhibit_ids = (
            existing_active_exhibit_ids - imported_existing_exhibit_ids - authoritative_exhibit_ids
        ) | imported_active_exhibit_ids
        if len(resulting_active_exhibit_ids) > MAX_ACTIVE_EXHIBITS:
            collision_issues.append(
                f"import would leave more than {MAX_ACTIVE_EXHIBITS} active exhibits "
                "in the database; explicitly include and deactivate obsolete exhibits first"
            )
        for row in dataset.halls:
            by_slug = hall_by_slug.get(row.slug)
            by_source = hall_by_source.get(row.source_record_id)
            by_name = hall_by_name.get(row.name)
            if by_slug is not None and by_slug.source_name not in {None, source_name}:
                collision_issues.append(f"hall slug '{row.slug}' is owned by source '{by_slug.source_name}'")
            if (
                by_slug is not None
                and by_slug.source_name == source_name
                and by_slug.source_record_id not in {None, row.source_record_id}
            ):
                collision_issues.append(
                    f"hall slug '{row.slug}' cannot change source_record_id "
                    f"from '{by_slug.source_record_id}' to '{row.source_record_id}'"
                )
            if by_source is not None and by_source.slug != row.slug:
                collision_issues.append(
                    f"hall source_record_id '{row.source_record_id}' cannot change slug "
                    f"from '{by_source.slug}' to '{row.slug}'"
                )
            if by_name is not None and by_name.slug != row.slug:
                collision_issues.append(f"hall name '{row.name}' is already used by slug '{by_name.slug}'")
        inactive_hall_slugs = {row.slug for row in dataset.halls if not row.is_active} | authoritative_hall_slugs
        if inactive_hall_slugs:
            planned_exhibits = {(source_name, row.source_record_id): row for row in dataset.exhibits}
            active_exhibits = (
                await self.session.execute(
                    select(Exhibit).where(
                        Exhibit.is_active.is_(True),
                        Exhibit.hall.in_(inactive_hall_slugs),
                    )
                )
            ).scalars()
            for exhibit in active_exhibits:
                if exhibit.id in authoritative_exhibit_ids:
                    continue
                planned = planned_exhibits.get((exhibit.source_name, exhibit.source_record_id))
                if planned is not None and (not planned.is_active or planned.hall not in inactive_hall_slugs):
                    continue
                collision_issues.append(
                    f"inactive hall '{exhibit.hall}' would retain active exhibit "
                    f"'{exhibit.source_record_id or exhibit.id}'; explicitly include "
                    "the exhibit as inactive or move it to an active hall"
                )
        if collision_issues:
            raise MuseumDataValidationError(collision_issues)

        hall_active_snapshots: dict[str, bool] = {model.slug: bool(model.is_active) for model in authoritative_halls}
        for row in dataset.halls:
            model = hall_by_slug.get(row.slug)
            was_active = bool(model and model.is_active)
            if model is not None:
                hall_active_snapshots[row.slug] = bool(model.is_active)
            values = {
                "name": row.name,
                "description": row.description,
                "floor": row.floor,
                "estimated_duration_minutes": row.estimated_duration_minutes,
                "display_order": row.display_order,
                "is_active": row.is_active,
                "suggested_questions": row.suggested_questions,
                "source_name": source_name,
                "source_record_id": row.source_record_id,
            }
            if model is None:
                model = Hall(slug=row.slug, **values)
                self.session.add(model)
                hall_by_slug[row.slug] = model
                summary.halls_created += 1
            elif _apply_changes(model, values):
                summary.halls_updated += 1
            else:
                summary.unchanged += 1
            if was_active and not row.is_active:
                summary.halls_deactivated += 1

        for model in authoritative_halls:
            if model.is_active:
                model.is_active = False
                summary.halls_updated += 1
                summary.halls_deactivated += 1

        to_index: list[tuple[Exhibit, ExhibitImportRow]] = []
        to_delete: list[Exhibit] = []
        restore_snapshots: dict[str, dict[str, Any]] = {}
        for row in dataset.exhibits:
            model = exhibit_by_source.get(row.source_record_id)
            indexed_values = {
                "name": row.name,
                "description": row.description,
                "location_x": row.location_x,
                "location_y": row.location_y,
                "floor": row.floor,
                "hall": row.hall,
                "category": row.category,
                "era": row.era,
                "importance": row.importance,
                "estimated_visit_time": row.estimated_visit_time,
                "display_order": row.display_order,
                "suggested_questions": row.suggested_questions,
                "source_name": source_name,
                "source_record_id": row.source_record_id,
            }
            values = dict(indexed_values)
            if row.image_url_present:
                values["image_url"] = row.image_url
            was_active = bool(model and model.is_active)
            if model is None:
                model = Exhibit(
                    id=deterministic_exhibit_id(source_name, row.source_record_id),
                    is_active=False,
                    **values,
                )
                self.session.add(model)
                exhibit_by_source[row.source_record_id] = model
                changed = True
                index_changed = True
                summary.exhibits_created += 1
            else:
                original_values = {key: getattr(model, key) for key in values}
                original_values["is_active"] = model.is_active
                index_changed = any(getattr(model, key) != value for key, value in indexed_values.items())
                changed = _apply_changes(model, values)
                if changed or model.is_active != row.is_active:
                    summary.exhibits_updated += 1
                else:
                    summary.unchanged += 1
                if was_active and (index_changed or not row.is_active):
                    restore_snapshots[model.id] = original_values
            if row.is_active and (index_changed or not was_active):
                model.is_active = False
                to_index.append((model, row))
                summary.pending_index.append(row.source_record_id)
            elif not row.is_active:
                model.is_active = False
                if was_active:
                    to_delete.append(model)
                    summary.exhibits_deactivated += 1

        for model in authoritative_exhibits:
            if not model.is_active:
                continue
            restore_snapshots[model.id] = {"is_active": True}
            model.is_active = False
            to_delete.append(model)
            summary.exhibits_updated += 1
            summary.exhibits_deactivated += 1

        if (to_index or to_delete) and self.indexing_service is None:
            await self.session.rollback()
            raise RuntimeError("RAG indexing service is required for active exhibit imports")

        failures: list[str] = []
        blocked_ids: set[str] = set()
        predeleted_ids: set[str] = set()
        restored_hall_slugs: set[str] = set()

        # Existing active records keep their previous DB state unless the old
        # RAG source is successfully removed. This avoids an inactive DB row
        # coexisting with a still-retrievable stale index after delete failure.
        for model in [*to_delete, *(item[0] for item in to_index if item[0].id in restore_snapshots)]:
            try:
                await self.indexing_service.delete_source(model.id, "exhibit")  # type: ignore[union-attr]
                # The legacy bootstrap indexed each placeholder Exhibit through
                # its linked Document rather than as an exhibit source. Remove
                # that vector source during authoritative takeover, but retain
                # the Document/IngestionJob database audit rows.
                if authoritative and model.source_name is None and model.document_id:
                    await self.indexing_service.delete_source(  # type: ignore[union-attr]
                        model.document_id,
                        "document",
                    )
                predeleted_ids.add(model.id)
            except Exception as exc:
                _apply_changes(model, restore_snapshots[model.id])
                restored_hall_slug = str(model.hall or "")
                restored_hall = hall_by_slug.get(restored_hall_slug)
                if (
                    restored_hall is not None
                    and restored_hall_slug in inactive_hall_slugs
                    and restored_hall_slug in hall_active_snapshots
                ):
                    original_hall_active = hall_active_snapshots[restored_hall_slug]
                    if (
                        original_hall_active
                        and not restored_hall.is_active
                        and restored_hall_slug not in restored_hall_slugs
                    ):
                        summary.halls_deactivated -= 1
                        restored_hall_slugs.add(restored_hall_slug)
                    restored_hall.is_active = original_hall_active
                blocked_ids.add(model.id)
                source_record_id = str(model.source_record_id or model.id)
                if source_record_id not in summary.pending_index:
                    summary.pending_index.append(source_record_id)
                failures.append(f"{source_record_id}: delete failed ({type(exc).__name__})")
                if model in to_delete:
                    summary.exhibits_deactivated -= 1
                    summary.exhibits_updated -= 1
                else:
                    summary.exhibits_updated -= 1

        to_index = [item for item in to_index if item[0].id not in blocked_ids]

        # New/changed records are now committed inactive before new chunks are
        # written. Only a completely indexed exhibit is activated below.
        await self.session.commit()

        for model, row in to_index:
            try:
                if model.id not in predeleted_ids:
                    await self.indexing_service.delete_source(model.id, "exhibit")  # type: ignore[union-attr]
                await self.indexing_service.index_source(_content_source(model, row))  # type: ignore[union-attr]
            except Exception as exc:  # external provider errors are summarized after all rows
                failures.append(f"{row.source_record_id}: index failed ({type(exc).__name__})")
                try:
                    await self.indexing_service.delete_source(model.id, "exhibit")  # type: ignore[union-attr]
                except Exception as cleanup_exc:
                    failures.append(
                        f"{row.source_record_id}: partial-index cleanup failed ({type(cleanup_exc).__name__})"
                    )
                continue
            model.is_active = True
            summary.exhibits_indexed += 1
            summary.pending_index.remove(row.source_record_id)

        await self.session.commit()
        if failures:
            raise MuseumDataIndexingError(failures, summary)
        return summary

    async def _existing_halls(
        self, dataset: MuseumDataset, source_name: str
    ) -> tuple[dict[str, Hall], dict[str, Hall], dict[str, Hall]]:
        slugs = [row.slug for row in dataset.halls]
        names = [row.name for row in dataset.halls]
        source_ids = [row.source_record_id for row in dataset.halls]
        by_slug: dict[str, Hall] = {}
        by_source: dict[str, Hall] = {}
        by_name: dict[str, Hall] = {}
        if slugs:
            rows = (await self.session.execute(select(Hall).where(Hall.slug.in_(slugs)))).scalars()
            by_slug.update({row.slug: row for row in rows})
        if source_ids:
            rows = (
                await self.session.execute(
                    select(Hall).where(
                        Hall.source_name == source_name,
                        Hall.source_record_id.in_(source_ids),
                    )
                )
            ).scalars()
            by_source.update({row.source_record_id: row for row in rows if row.source_record_id})
        if names:
            rows = (await self.session.execute(select(Hall).where(Hall.name.in_(names)))).scalars()
            by_name.update({row.name: row for row in rows})
        return by_slug, by_source, by_name

    async def _existing_exhibits(self, dataset: MuseumDataset, source_name: str) -> dict[str, Exhibit]:
        source_ids = [row.source_record_id for row in dataset.exhibits]
        if not source_ids:
            return {}
        rows = (
            await self.session.execute(
                select(Exhibit).where(
                    Exhibit.source_name == source_name,
                    Exhibit.source_record_id.in_(source_ids),
                )
            )
        ).scalars()
        return {row.source_record_id: row for row in rows if row.source_record_id}

    async def _authoritative_targets(
        self,
        dataset: MuseumDataset,
        source_name: str,
    ) -> tuple[list[Hall], list[Exhibit]]:
        """Find omitted same-source and unowned legacy rows eligible for takeover."""
        incoming_hall_slugs = {row.slug for row in dataset.halls}
        incoming_hall_source_ids = {row.source_record_id for row in dataset.halls}
        hall_rows = (
            await self.session.execute(
                select(Hall).where(or_(Hall.source_name.is_(None), Hall.source_name == source_name))
            )
        ).scalars()
        halls = [
            model
            for model in hall_rows
            if model.slug not in incoming_hall_slugs
            and not (model.source_name == source_name and model.source_record_id in incoming_hall_source_ids)
        ]

        incoming_exhibit_source_ids = {row.source_record_id for row in dataset.exhibits}
        exhibit_rows = (
            await self.session.execute(
                select(Exhibit).where(
                    or_(
                        Exhibit.source_name.is_(None),
                        Exhibit.source_name == source_name,
                    )
                )
            )
        ).scalars()
        exhibits = [
            model
            for model in exhibit_rows
            if not (model.source_name == source_name and model.source_record_id in incoming_exhibit_source_ids)
        ]
        return halls, exhibits


def _apply_changes(model: Any, values: dict[str, Any]) -> bool:
    changed = False
    for key, value in values.items():
        if getattr(model, key) != value:
            setattr(model, key, value)
            changed = True
    return changed


def _content_source(model: Exhibit, row: ExhibitImportRow) -> ContentSource:
    content = "\n".join(
        part
        for part in (
            f"展品名称：{row.name}",
            f"所在展厅：{row.hall}",
            f"年代：{row.era}" if row.era else "",
            f"类别：{row.category}" if row.category else "",
            f"展品介绍：{row.description}",
        )
        if part
    )
    return ContentSource(
        source_id=model.id,
        source_type="exhibit",
        content=content,
        metadata=ContentMetadata(
            name=row.name,
            category=row.category,
            hall=row.hall,
            floor=row.floor,
            era=row.era,
            importance=row.importance,
            location_x=row.location_x,
            location_y=row.location_y,
            extra={
                "exhibit_id": model.id,
                "source_name": model.source_name,
                "source_record_id": model.source_record_id,
                "suggested_questions": row.suggested_questions,
            },
        ),
    )
