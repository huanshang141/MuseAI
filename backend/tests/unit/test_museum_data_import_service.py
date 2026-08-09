import csv
from dataclasses import replace
from pathlib import Path
from zipfile import ZipFile

import pytest
from app.application.museum_data_import_service import (
    EXHIBIT_HEADERS,
    EXHIBIT_OPTIONAL_HEADERS,
    HALL_HEADERS,
    MAX_ACTIVE_EXHIBITS,
    SUGGESTION_MAX_LENGTH,
    ExhibitImportRow,
    HallImportRow,
    MuseumDataImportService,
    MuseumDataIndexingError,
    MuseumDataset,
    MuseumDataValidationError,
    deterministic_exhibit_id,
    load_museum_dataset,
)
from app.infra.postgres.models import Base, Exhibit, Hall
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine


def _hall_values(**overrides):
    values = {
        "source_record_id": "hall-001",
        "slug": "basic-hall",
        "name": "基本展厅",
        "description": "可信的展厅介绍",
        "floor": 1,
        "estimated_duration_minutes": 30,
        "display_order": 1,
        "is_active": True,
        "suggested_questions": "这里最值得看什么？|这个展厅讲什么？",
    }
    values.update(overrides)
    return values


def _exhibit_values(**overrides):
    values = {
        "source_record_id": "exhibit-001",
        "name": "人面鱼纹彩陶盆",
        "description": "真实展品介绍",
        "hall": "basic-hall",
        "floor": 1,
        "category": "彩陶",
        "era": "新石器时代",
        "importance": 90,
        "estimated_visit_time": 300,
        "display_order": 1,
        "location_x": 10.5,
        "location_y": 20.25,
        "is_active": True,
        "suggested_questions": '["纹样能说明什么？"]',
    }
    values.update(overrides)
    return values


def _write_csv(path: Path, headers, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _dataset(*, exhibit_description="真实展品介绍", exhibit_active=True):
    return MuseumDataset(
        halls=[
            HallImportRow(
                source_record_id="hall-001",
                slug="basic-hall",
                name="基本展厅",
                description="可信的展厅介绍",
                floor=1,
                estimated_duration_minutes=30,
                display_order=1,
                is_active=True,
                suggested_questions=["这里最值得看什么？"],
            )
        ],
        exhibits=[
            ExhibitImportRow(
                source_record_id="exhibit-001",
                name="人面鱼纹彩陶盆",
                description=exhibit_description,
                hall="basic-hall",
                floor=1,
                category="彩陶",
                era="新石器时代",
                importance=90,
                estimated_visit_time=300,
                display_order=1,
                location_x=10.5,
                location_y=20.25,
                is_active=exhibit_active,
                suggested_questions=["纹样能说明什么？"],
            )
        ],
    )


class FakeIndexer:
    def __init__(
        self,
        *,
        fail=False,
        fail_delete=False,
        fail_delete_ids: set[str] | None = None,
    ):
        self.fail = fail
        self.fail_delete = fail_delete
        self.fail_delete_ids = fail_delete_ids or set()
        self.indexed = []
        self.deleted = []

    async def delete_source(self, source_id, source_type=None):
        if self.fail_delete or source_id in self.fail_delete_ids:
            raise RuntimeError("elasticsearch delete unavailable")
        self.deleted.append((source_id, source_type))
        return {"deleted": 1}

    async def index_source(self, source, max_concurrency=10):
        if self.fail:
            raise RuntimeError("embedding service unavailable")
        self.indexed.append(source)
        return 1


@pytest.fixture
async def import_session():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    maker = async_sessionmaker(engine, expire_on_commit=False)
    async with maker() as session:
        yield session
    await engine.dispose()


def test_loads_utf8_sig_csv_pair_and_pipe_or_json_questions(tmp_path):
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    _write_csv(tmp_path / "exhibits.csv", EXHIBIT_HEADERS, [_exhibit_values()])

    dataset = load_museum_dataset(tmp_path)

    assert dataset.halls[0].name == "基本展厅"
    assert dataset.halls[0].suggested_questions == ["这里最值得看什么？", "这个展厅讲什么？"]
    assert dataset.exhibits[0].suggested_questions == ["纹样能说明什么？"]


def test_rejects_suggestion_over_runtime_length_contract(tmp_path):
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    _write_csv(
        tmp_path / "exhibits.csv",
        EXHIBIT_HEADERS,
        [_exhibit_values(suggested_questions="问" * (SUGGESTION_MAX_LENGTH + 1))],
    )

    with pytest.raises(
        MuseumDataValidationError,
        match=rf"exceeds {SUGGESTION_MAX_LENGTH} characters",
    ):
        load_museum_dataset(tmp_path)


def test_csv_and_xlsx_accept_optional_https_image_url(tmp_path):
    exhibit_headers = (*EXHIBIT_HEADERS, *EXHIBIT_OPTIONAL_HEADERS)
    image_url = "https://museum.example/images/pottery.png?version=1"
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    _write_csv(
        tmp_path / "exhibits.csv",
        exhibit_headers,
        [_exhibit_values(image_url=image_url)],
    )

    csv_dataset = load_museum_dataset(tmp_path)

    assert csv_dataset.exhibits[0].image_url == image_url
    assert csv_dataset.exhibits[0].image_url_present is True

    workbook_path = tmp_path / "museum_data.xlsx"
    workbook = Workbook()
    halls = workbook.active
    halls.title = "halls"
    halls.append(HALL_HEADERS)
    halls.append([_hall_values()[header] for header in HALL_HEADERS])
    exhibits = workbook.create_sheet("exhibits")
    exhibits.append(exhibit_headers)
    values = _exhibit_values(image_url=image_url)
    exhibits.append([values[header] for header in exhibit_headers])
    workbook.save(workbook_path)

    xlsx_dataset = load_museum_dataset(workbook_path)
    assert xlsx_dataset.exhibits[0].image_url == image_url
    assert xlsx_dataset.exhibits[0].image_url_present is True


def test_import_rejects_non_https_image_url(tmp_path):
    exhibit_headers = (*EXHIBIT_HEADERS, *EXHIBIT_OPTIONAL_HEADERS)
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    _write_csv(
        tmp_path / "exhibits.csv",
        exhibit_headers,
        [_exhibit_values(image_url="http://museum.example/pottery.png")],
    )

    with pytest.raises(MuseumDataValidationError, match="absolute HTTPS URL"):
        load_museum_dataset(tmp_path)


def test_versioned_museum_template_has_trusted_nine_halls_and_no_fake_exhibits():
    template_dir = Path(__file__).resolve().parents[3] / "data" / "museum_template"
    with (template_dir / "exhibits.csv").open(encoding="utf-8-sig", newline="") as csv_file:
        assert "image_url" in (csv.DictReader(csv_file).fieldnames or [])

    dataset = load_museum_dataset(template_dir)

    assert [hall.slug for hall in dataset.halls] == [
        "basic-exhibition-hall",
        "site-protection-hall",
        "kiln-hall",
        "prehistoric-workshop",
        "banpo-girl-sculpture",
        "education-center",
        "peony-garden",
        "temporary-hall-1",
        "temporary-hall-2",
    ]
    assert [hall.display_order for hall in dataset.halls] == list(range(10, 100, 10))
    assert all(hall.name and hall.description and hall.is_active for hall in dataset.halls)
    assert all(hall.floor is None for hall in dataset.halls)
    assert all(hall.estimated_duration_minutes == 0 for hall in dataset.halls)
    assert dataset.exhibits == []


def test_versioned_test_dataset_is_explicit_and_covers_all_nine_halls():
    dataset_dir = Path(__file__).resolve().parents[3] / "data" / "museum_test_data"
    with (dataset_dir / "exhibits.csv").open(encoding="utf-8-sig", newline="") as csv_file:
        assert "image_url" in (csv.DictReader(csv_file).fieldnames or [])

    dataset = load_museum_dataset(dataset_dir)

    assert len(dataset.halls) == 9
    assert len(dataset.exhibits) == 46
    assert {exhibit.hall for exhibit in dataset.exhibits} == {hall.slug for hall in dataset.halls}
    assert all(exhibit.source_record_id.startswith("legacy-test-") for exhibit in dataset.exhibits)
    assert all(exhibit.name.startswith("【测试】") for exhibit in dataset.exhibits)
    assert all(exhibit.description.startswith("【测试数据，非馆方真实展品信息】") for exhibit in dataset.exhibits)
    assert all(exhibit.category == "测试数据" for exhibit in dataset.exhibits)
    assert all(exhibit.floor is None and exhibit.era is None for exhibit in dataset.exhibits)
    assert all(exhibit.image_url is None and exhibit.image_url_present for exhibit in dataset.exhibits)


def test_rejects_non_utf8_csv_as_structured_validation_error(tmp_path):
    (tmp_path / "halls.csv").write_bytes(b"\xff\xfe\x00\x00")
    _write_csv(tmp_path / "exhibits.csv", EXHIBIT_HEADERS, [_exhibit_values()])

    with pytest.raises(MuseumDataValidationError, match="must be UTF-8 encoded"):
        load_museum_dataset(tmp_path)


def test_rejects_malformed_csv_as_structured_validation_error(tmp_path):
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    values = _exhibit_values()
    row_prefix = ",".join(str(values[header]) for header in EXHIBIT_HEADERS[:-1])
    malformed = f'{",".join(EXHIBIT_HEADERS)}\n{row_prefix},"unterminated'
    (tmp_path / "exhibits.csv").write_text(malformed, encoding="utf-8")

    with pytest.raises(MuseumDataValidationError, match="is malformed near line 2"):
        load_museum_dataset(tmp_path)


def test_loads_exact_two_sheet_xlsx(tmp_path):
    path = tmp_path / "museum_data.xlsx"
    workbook = Workbook()
    halls = workbook.active
    halls.title = "halls"
    halls.append(HALL_HEADERS)
    halls.append([_hall_values()[header] for header in HALL_HEADERS])
    exhibits = workbook.create_sheet("exhibits")
    exhibits.append(EXHIBIT_HEADERS)
    exhibits.append([_exhibit_values()[header] for header in EXHIBIT_HEADERS])
    workbook.save(path)

    dataset = load_museum_dataset(path)

    assert len(dataset.halls) == 1
    assert dataset.exhibits[0].name == "人面鱼纹彩陶盆"


def test_rejects_corrupt_xlsx_as_structured_validation_error(tmp_path):
    path = tmp_path / "museum_data.xlsx"
    path.write_bytes(b"not-a-valid-xlsx")

    with pytest.raises(MuseumDataValidationError, match="not a valid .xlsx file"):
        load_museum_dataset(path)


def test_rejects_xlsx_zip_missing_content_types_as_validation_error(tmp_path):
    path = tmp_path / "museum_data.xlsx"
    with ZipFile(path, "w"):
        pass

    with pytest.raises(MuseumDataValidationError, match="not a valid .xlsx file"):
        load_museum_dataset(path)


def test_rejects_malformed_xlsx_xml_as_validation_error(tmp_path):
    path = tmp_path / "museum_data.xlsx"
    with ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types>")

    with pytest.raises(MuseumDataValidationError, match="not a valid .xlsx file"):
        load_museum_dataset(path)


def test_wraps_explicit_openpyxl_value_error_but_not_os_errors(tmp_path, monkeypatch):
    path = tmp_path / "museum_data.xlsx"
    path.write_bytes(b"placeholder")
    monkeypatch.setattr(
        "app.application.museum_data_import_service.load_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("invalid metadata")),
    )
    with pytest.raises(MuseumDataValidationError, match="invalid metadata"):
        load_museum_dataset(path)

    monkeypatch.setattr(
        "app.application.museum_data_import_service.load_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(PermissionError("access denied")),
    )
    with pytest.raises(PermissionError, match="access denied"):
        load_museum_dataset(path)


@pytest.mark.parametrize("problem", ["missing_header", "unknown_hall", "duplicate_key", "duplicate_hall_name"])
def test_rejects_invalid_csv_before_import(tmp_path, problem):
    hall_headers = list(HALL_HEADERS)
    hall_rows = [_hall_values()]
    exhibit_rows = [_exhibit_values()]
    if problem == "missing_header":
        hall_headers.remove("description")
    elif problem == "unknown_hall":
        exhibit_rows[0]["hall"] = "not-imported"
    elif problem == "duplicate_key":
        exhibit_rows.append(_exhibit_values(name="重复稳定键"))
    else:
        hall_rows.append(_hall_values(source_record_id="hall-002", slug="second-hall"))
    _write_csv(tmp_path / "halls.csv", hall_headers, hall_rows)
    _write_csv(tmp_path / "exhibits.csv", EXHIBIT_HEADERS, exhibit_rows)

    with pytest.raises(MuseumDataValidationError):
        load_museum_dataset(tmp_path)


def test_rejects_extra_xlsx_sheet(tmp_path):
    path = tmp_path / "museum_data.xlsx"
    workbook = Workbook()
    workbook.active.title = "halls"
    workbook["halls"].append(HALL_HEADERS)
    workbook["halls"].append([_hall_values()[header] for header in HALL_HEADERS])
    workbook.create_sheet("exhibits").append(EXHIBIT_HEADERS)
    workbook["exhibits"].append([_exhibit_values()[header] for header in EXHIBIT_HEADERS])
    workbook.create_sheet("notes")
    workbook.save(path)

    with pytest.raises(MuseumDataValidationError, match="unexpected sheets"):
        load_museum_dataset(path)


def test_rejects_more_than_nine_active_halls(tmp_path):
    hall_rows = [
        _hall_values(
            source_record_id=f"hall-{index}",
            slug=f"hall-{index}",
            name=f"展厅{index}",
            display_order=index,
        )
        for index in range(10)
    ]
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, hall_rows)
    _write_csv(tmp_path / "exhibits.csv", EXHIBIT_HEADERS, [])

    with pytest.raises(MuseumDataValidationError, match="at most 9 active halls"):
        load_museum_dataset(tmp_path)


def test_rejects_more_than_two_thousand_active_exhibits(tmp_path):
    _write_csv(tmp_path / "halls.csv", HALL_HEADERS, [_hall_values()])
    exhibit_rows = [
        _exhibit_values(
            source_record_id=f"exhibit-{index}",
            name=f"真实展品{index}",
            display_order=index,
        )
        for index in range(MAX_ACTIVE_EXHIBITS + 1)
    ]
    _write_csv(tmp_path / "exhibits.csv", EXHIBIT_HEADERS, exhibit_rows)

    with pytest.raises(MuseumDataValidationError, match="at most 2000 active exhibits"):
        load_museum_dataset(tmp_path)


@pytest.mark.asyncio
async def test_dry_run_needs_no_database_or_indexer():
    summary = await MuseumDataImportService().import_dataset(_dataset(), source_name="banpo-2026", dry_run=True)

    assert summary.dry_run is True
    assert summary.pending_index == ["exhibit-001"]

    inactive = _dataset(exhibit_active=False)
    authoritative = await MuseumDataImportService().import_dataset(
        MuseumDataset(
            halls=[replace(inactive.halls[0], is_active=False)],
            exhibits=inactive.exhibits,
        ),
        source_name="banpo-2026",
        dry_run=True,
        authoritative=True,
    )
    assert authoritative.authoritative is True
    assert authoritative.authoritative_cleanup_deferred is True
    assert authoritative.halls_planned_deactivation == 1
    assert authoritative.exhibits_planned_deactivation == 1


@pytest.mark.asyncio
async def test_idempotent_upsert_indexes_only_new_or_changed(import_session):
    indexer = FakeIndexer()
    service = MuseumDataImportService(import_session, indexer)

    first = await service.import_dataset(_dataset(), source_name="banpo-2026")
    second = await service.import_dataset(_dataset(), source_name="banpo-2026")
    changed = await service.import_dataset(_dataset(exhibit_description="更新后的真实介绍"), source_name="banpo-2026")

    assert first.halls_created == 1
    assert first.exhibits_created == 1
    assert first.exhibits_indexed == 1
    assert second.exhibits_indexed == 0
    assert changed.exhibits_updated == 1
    assert changed.exhibits_indexed == 1
    assert len(indexer.indexed) == 2
    assert (await import_session.scalar(select(func.count()).select_from(Hall))) == 1
    assert (await import_session.scalar(select(func.count()).select_from(Exhibit))) == 1
    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    assert exhibit.description == "更新后的真实介绍"
    assert exhibit.is_active is True


@pytest.mark.asyncio
async def test_imported_image_url_updates_without_rebuilding_rag(import_session):
    indexer = FakeIndexer()
    service = MuseumDataImportService(import_session, indexer)
    dataset = _dataset()
    await service.import_dataset(dataset, source_name="banpo-2026")
    with_image = MuseumDataset(
        halls=dataset.halls,
        exhibits=[
            replace(
                dataset.exhibits[0],
                image_url="https://museum.example/pottery.png",
                image_url_present=True,
            )
        ],
    )

    summary = await service.import_dataset(with_image, source_name="banpo-2026")

    exhibit = await import_session.get(
        Exhibit,
        deterministic_exhibit_id("banpo-2026", "exhibit-001"),
    )
    assert exhibit.image_url == "https://museum.example/pottery.png"
    assert exhibit.is_active is True
    assert summary.exhibits_updated == 1
    assert summary.exhibits_indexed == 0
    assert len(indexer.indexed) == 1


@pytest.mark.asyncio
async def test_existing_hall_name_collision_is_structured_before_mutation(import_session):
    import_session.add(
        Hall(
            slug="legacy-hall",
            name="基本展厅",
            description="旧数据",
            estimated_duration_minutes=20,
            display_order=1,
            is_active=True,
        )
    )
    await import_session.commit()

    with pytest.raises(MuseumDataValidationError, match="already used by slug"):
        await MuseumDataImportService(import_session, FakeIndexer()).import_dataset(
            _dataset(), source_name="banpo-2026"
        )

    assert await import_session.get(Hall, "basic-hall") is None


@pytest.mark.asyncio
async def test_existing_hall_source_record_id_cannot_change_before_mutation(
    import_session,
):
    existing = Hall(
        slug="basic-hall",
        name="基本展厅",
        description="导入前介绍",
        estimated_duration_minutes=20,
        display_order=1,
        is_active=True,
        source_name="banpo-2026",
        source_record_id="hall-original",
    )
    import_session.add(existing)
    await import_session.commit()

    with pytest.raises(MuseumDataValidationError, match="cannot change source_record_id"):
        await MuseumDataImportService(import_session, FakeIndexer()).import_dataset(
            _dataset(), source_name="banpo-2026"
        )

    await import_session.refresh(existing)
    assert existing.source_record_id == "hall-original"
    assert existing.description == "导入前介绍"
    assert await import_session.scalar(select(func.count()).select_from(Exhibit)) == 0


@pytest.mark.asyncio
async def test_import_rejects_more_than_nine_resulting_database_halls(import_session):
    import_session.add_all(
        [
            Hall(
                slug=f"legacy-{index}",
                name=f"旧展厅{index}",
                description="旧数据",
                estimated_duration_minutes=20,
                display_order=index,
                is_active=True,
            )
            for index in range(9)
        ]
    )
    await import_session.commit()

    with pytest.raises(MuseumDataValidationError, match="more than 9 active halls"):
        await MuseumDataImportService(import_session, FakeIndexer()).import_dataset(
            _dataset(), source_name="banpo-2026"
        )

    assert await import_session.get(Hall, "basic-hall") is None


@pytest.mark.asyncio
async def test_import_rejects_more_than_two_thousand_resulting_active_exhibits(
    import_session,
):
    import_session.add_all(
        [
            Exhibit(
                id=f"legacy-{index}",
                name=f"旧展品{index}",
                hall="legacy-hall",
                is_active=True,
            )
            for index in range(MAX_ACTIVE_EXHIBITS)
        ]
    )
    await import_session.commit()

    with pytest.raises(
        MuseumDataValidationError,
        match="more than 2000 active exhibits in the database",
    ):
        await MuseumDataImportService(import_session, FakeIndexer()).import_dataset(
            _dataset(), source_name="banpo-2026"
        )

    assert await import_session.get(Hall, "basic-hall") is None
    assert (
        await import_session.scalar(select(func.count()).select_from(Exhibit).where(Exhibit.is_active.is_(True)))
    ) == MAX_ACTIVE_EXHIBITS


@pytest.mark.asyncio
async def test_omitted_rows_are_not_deleted(import_session):
    service = MuseumDataImportService(import_session, FakeIndexer())
    await service.import_dataset(_dataset(), source_name="banpo-2026")
    halls_only = MuseumDataset(halls=_dataset().halls, exhibits=[])

    await service.import_dataset(halls_only, source_name="banpo-2026")

    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    assert exhibit is not None
    assert exhibit.is_active is True


@pytest.mark.asyncio
async def test_default_import_keeps_legacy_placeholders_active(import_session):
    legacy_hall = Hall(
        slug="legacy-placeholder-hall",
        name="旧占位展厅",
        description="旧占位数据",
        estimated_duration_minutes=20,
        display_order=90,
        is_active=True,
    )
    legacy_exhibit = Exhibit(
        id="legacy-placeholder-exhibit",
        name="旧占位展品",
        description="旧占位数据",
        hall=legacy_hall.slug,
        is_active=True,
    )
    import_session.add_all([legacy_hall, legacy_exhibit])
    await import_session.commit()
    indexer = FakeIndexer()

    summary = await MuseumDataImportService(
        import_session,
        indexer,
    ).import_dataset(_dataset(), source_name="banpo-2026")

    await import_session.refresh(legacy_hall)
    await import_session.refresh(legacy_exhibit)
    assert summary.authoritative is False
    assert legacy_hall.is_active is True
    assert legacy_exhibit.is_active is True
    assert (legacy_exhibit.id, "exhibit") not in indexer.deleted


@pytest.mark.asyncio
async def test_authoritative_import_deactivates_omitted_and_legacy_only(
    import_session,
):
    same_hall = Hall(
        slug="same-source-old-hall",
        name="同源旧展厅",
        description="同源遗漏数据",
        estimated_duration_minutes=20,
        display_order=80,
        is_active=True,
        source_name="banpo-2026",
        source_record_id="hall-old",
    )
    legacy_hall = Hall(
        slug="legacy-old-hall",
        name="无来源旧展厅",
        description="首次接管占位数据",
        estimated_duration_minutes=20,
        display_order=81,
        is_active=True,
    )
    other_hall = Hall(
        slug="other-source-hall",
        name="其他来源展厅",
        description="不得修改",
        estimated_duration_minutes=20,
        display_order=82,
        is_active=True,
        source_name="other-source",
        source_record_id="other-hall",
    )
    same_exhibit = Exhibit(
        id="same-source-old-exhibit",
        name="同源旧展品",
        hall=same_hall.slug,
        is_active=True,
        source_name="banpo-2026",
        source_record_id="exhibit-old",
        image_path="same-source-old-exhibit/upload.png",
    )
    legacy_exhibit = Exhibit(
        id="legacy-old-exhibit",
        name="无来源旧展品",
        hall=legacy_hall.slug,
        document_id="legacy-old-document",
        is_active=True,
        image_path="legacy-old-exhibit/upload.png",
    )
    other_exhibit = Exhibit(
        id="other-source-exhibit",
        name="其他来源展品",
        hall=other_hall.slug,
        is_active=True,
        source_name="other-source",
        source_record_id="other-exhibit",
    )
    import_session.add_all(
        [
            same_hall,
            legacy_hall,
            other_hall,
            same_exhibit,
            legacy_exhibit,
            other_exhibit,
        ]
    )
    await import_session.commit()
    indexer = FakeIndexer()

    summary = await MuseumDataImportService(
        import_session,
        indexer,
    ).import_dataset(
        _dataset(),
        source_name="banpo-2026",
        authoritative=True,
    )

    for model in (same_hall, legacy_hall, other_hall, same_exhibit, legacy_exhibit, other_exhibit):
        await import_session.refresh(model)
    assert same_hall.is_active is False
    assert legacy_hall.is_active is False
    assert same_exhibit.is_active is False
    assert legacy_exhibit.is_active is False
    # Deactivation preserves uploaded files and their references so a rollback
    # or later reactivation can restore the exhibit without re-uploading.
    assert same_exhibit.image_path == "same-source-old-exhibit/upload.png"
    assert legacy_exhibit.image_path == "legacy-old-exhibit/upload.png"
    assert other_hall.is_active is True
    assert other_exhibit.is_active is True
    assert summary.authoritative is True
    assert summary.halls_planned_deactivation == 2
    assert summary.exhibits_planned_deactivation == 2
    assert summary.halls_deactivated == 2
    assert summary.exhibits_deactivated == 2
    assert (same_exhibit.id, "exhibit") in indexer.deleted
    assert (legacy_exhibit.id, "exhibit") in indexer.deleted
    assert (legacy_exhibit.document_id, "document") in indexer.deleted
    assert (other_exhibit.id, "exhibit") not in indexer.deleted


@pytest.mark.asyncio
async def test_authoritative_delete_failure_keeps_failed_group_active(
    import_session,
):
    failed_hall = Hall(
        slug="legacy-failed-hall",
        name="删除失败旧展厅",
        description="应整体保留",
        estimated_duration_minutes=20,
        display_order=80,
        is_active=True,
    )
    successful_hall = Hall(
        slug="same-source-success-hall",
        name="删除成功旧展厅",
        description="应正常停用",
        estimated_duration_minutes=20,
        display_order=81,
        is_active=True,
        source_name="banpo-2026",
        source_record_id="hall-success-old",
    )
    failed_exhibit = Exhibit(
        id="legacy-failed-exhibit",
        name="删除失败旧展品",
        hall=failed_hall.slug,
        document_id="legacy-failed-document",
        is_active=True,
    )
    successful_exhibit = Exhibit(
        id="same-source-success-exhibit",
        name="删除成功旧展品",
        hall=successful_hall.slug,
        is_active=True,
        source_name="banpo-2026",
        source_record_id="exhibit-success-old",
    )
    import_session.add_all([failed_hall, successful_hall, failed_exhibit, successful_exhibit])
    await import_session.commit()
    indexer = FakeIndexer(fail_delete_ids={failed_exhibit.document_id})

    with pytest.raises(MuseumDataIndexingError) as captured:
        await MuseumDataImportService(import_session, indexer).import_dataset(
            _dataset(),
            source_name="banpo-2026",
            authoritative=True,
        )

    for model in (failed_hall, successful_hall, failed_exhibit, successful_exhibit):
        await import_session.refresh(model)
    assert failed_hall.is_active is True
    assert failed_exhibit.is_active is True
    assert successful_hall.is_active is False
    assert successful_exhibit.is_active is False
    assert failed_exhibit.id in captured.value.summary.pending_index
    assert captured.value.summary.halls_deactivated == 1
    assert captured.value.summary.exhibits_deactivated == 1
    assert (failed_exhibit.id, "exhibit") in indexer.deleted
    assert (failed_exhibit.document_id, "document") not in indexer.deleted


@pytest.mark.asyncio
async def test_inactive_hall_requires_explicit_exhibit_resolution(import_session):
    service = MuseumDataImportService(import_session, FakeIndexer())
    original = _dataset()
    await service.import_dataset(original, source_name="banpo-2026")
    inactive_hall = replace(original.halls[0], is_active=False)

    with pytest.raises(MuseumDataValidationError, match="would retain active exhibit"):
        await service.import_dataset(
            MuseumDataset(halls=[inactive_hall], exhibits=[]),
            source_name="banpo-2026",
        )

    hall = await import_session.get(Hall, "basic-hall")
    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    assert hall.is_active is True
    assert exhibit.is_active is True


@pytest.mark.asyncio
async def test_inactive_hall_and_explicit_inactive_exhibit_are_consistent(import_session):
    service = MuseumDataImportService(import_session, FakeIndexer())
    original = _dataset()
    await service.import_dataset(original, source_name="banpo-2026")

    await service.import_dataset(
        MuseumDataset(
            halls=[replace(original.halls[0], is_active=False)],
            exhibits=[replace(original.exhibits[0], is_active=False)],
        ),
        source_name="banpo-2026",
    )

    hall = await import_session.get(Hall, "basic-hall")
    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    assert hall.is_active is False
    assert exhibit.is_active is False


@pytest.mark.asyncio
async def test_index_failure_leaves_exhibit_inactive_and_reports_pending(import_session):
    service = MuseumDataImportService(import_session, FakeIndexer(fail=True))

    with pytest.raises(MuseumDataIndexingError) as captured:
        await service.import_dataset(_dataset(), source_name="banpo-2026")

    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    await import_session.refresh(exhibit)
    assert exhibit.is_active is False
    assert captured.value.summary.pending_index == ["exhibit-001"]


@pytest.mark.asyncio
async def test_explicit_deactivation_removes_rag_source(import_session):
    indexer = FakeIndexer()
    service = MuseumDataImportService(import_session, indexer)
    await service.import_dataset(_dataset(), source_name="banpo-2026")

    summary = await service.import_dataset(_dataset(exhibit_active=False), source_name="banpo-2026")

    exhibit_id = deterministic_exhibit_id("banpo-2026", "exhibit-001")
    exhibit = await import_session.get(Exhibit, exhibit_id)
    assert exhibit.is_active is False
    assert summary.exhibits_deactivated == 1
    assert indexer.deleted[-1] == (exhibit_id, "exhibit")


@pytest.mark.asyncio
async def test_delete_failure_keeps_previously_active_exhibit_visible_for_retry(import_session):
    healthy_indexer = FakeIndexer()
    await MuseumDataImportService(import_session, healthy_indexer).import_dataset(_dataset(), source_name="banpo-2026")
    failing_service = MuseumDataImportService(import_session, FakeIndexer(fail_delete=True))

    with pytest.raises(MuseumDataIndexingError) as captured:
        await failing_service.import_dataset(_dataset(exhibit_active=False), source_name="banpo-2026")

    exhibit = await import_session.get(Exhibit, deterministic_exhibit_id("banpo-2026", "exhibit-001"))
    await import_session.refresh(exhibit)
    assert exhibit.is_active is True
    assert captured.value.summary.pending_index == ["exhibit-001"]


@pytest.mark.asyncio
async def test_hall_deactivation_restores_only_hall_with_failed_exhibit_delete(
    import_session,
):
    first = _dataset()
    active_dataset = MuseumDataset(
        halls=[
            first.halls[0],
            replace(
                first.halls[0],
                source_record_id="hall-002",
                slug="site-hall",
                name="遗址展厅",
                display_order=2,
            ),
        ],
        exhibits=[
            first.exhibits[0],
            replace(
                first.exhibits[0],
                source_record_id="exhibit-002",
                name="遗址展品",
                hall="site-hall",
                display_order=2,
            ),
        ],
    )
    await MuseumDataImportService(import_session, FakeIndexer()).import_dataset(
        active_dataset,
        source_name="banpo-2026",
    )

    failed_exhibit_id = deterministic_exhibit_id("banpo-2026", "exhibit-001")
    successful_exhibit_id = deterministic_exhibit_id("banpo-2026", "exhibit-002")
    failing_indexer = FakeIndexer(fail_delete_ids={failed_exhibit_id})
    inactive_dataset = MuseumDataset(
        halls=[replace(hall, is_active=False) for hall in active_dataset.halls],
        exhibits=[replace(exhibit, is_active=False) for exhibit in active_dataset.exhibits],
    )

    with pytest.raises(MuseumDataIndexingError) as captured:
        await MuseumDataImportService(
            import_session,
            failing_indexer,
        ).import_dataset(inactive_dataset, source_name="banpo-2026")

    failed_hall = await import_session.get(Hall, "basic-hall")
    successful_hall = await import_session.get(Hall, "site-hall")
    failed_exhibit = await import_session.get(Exhibit, failed_exhibit_id)
    successful_exhibit = await import_session.get(Exhibit, successful_exhibit_id)
    assert failed_hall.is_active is True
    assert failed_exhibit.is_active is True
    assert successful_hall.is_active is False
    assert successful_exhibit.is_active is False
    assert failing_indexer.deleted == [(successful_exhibit_id, "exhibit")]
    assert captured.value.summary.pending_index == ["exhibit-001"]
    assert captured.value.summary.exhibits_deactivated == 1
